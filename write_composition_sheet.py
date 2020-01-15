import read_model_ordernum as rd
import pandas as pd
import openpyxl
import pprint
import time

#df=rd.format_list()
#print(type(df))

#path='testfile/test/【マスター簡易】機器構成シート_型番照合_20191227.xls'
#path2='【マスター簡易】機器構成シート_型番照合_20191227.xlsx'
def_path = 'SystemAutomation.xlsx'
df_def=pd.read_excel(def_path,sheet_name='HPE_Server',header=0,index_col=0)
print(df_def)
print("定義ファイルを読み込みました")

    
input_wb=rd.select_file()
print(input_wb)
wb = openpyxl.load_workbook(input_wb)
print("書き込みファイルを読み込みました")

print(type(wb))
# <class 'openpyxl.workbook.workbook.Workbook'>

print(wb.sheetnames)
# ['sheet1', 'sheet2']

#HPE用のシートの取得
sheet = wb['HPE用']

#サーバー本体の型番のあるセルを取得
cell=sheet['B12']
#セルの型と値を正しくセルが選択されたかテストで出力
print(type(cell))
print(cell.value)

#構成シートのサーバー本体の型番と定義表の型番を比較
for i in range(0,7):
    if cell.value==df_def.iat[i,0]:
        #型番が合うものが定義ファイルに見つかったらMatch!と出力
        print("Match!")
        #一部テストで出力してみる
        print(df_def.iat[i,3])
        print(df_def.iat[i,2])
        print(df_def.iat[i,1])
        #構成シートの各セルに記入
        #標準メモリを記入
        sheet.cell(row=12, column=41, value=df_def.iat[i,2])
        #RAIDコントローラー名を記入
        sheet.cell(row=13, column=41, value=df_def.iat[i,3])
        #RAID FW versionを記入
        sheet.cell(row=36, column=24, value=df_def.iat[i,4])
        #サーバー本体のHDDサイズを記入
        sheet.cell(row=14, column=41, value=df_def.iat[i,5])
        sheet.cell(row=15, column=41, value=df_def.iat[i,6])
        sheet.cell(row=16, column=41, value=df_def.iat[i,7])
        #NIC versionを記入
        sheet.cell(row=37, column=24, value=df_def.iat[i,8])
        
        #電源ユニットを記入
        sheet.cell(row=17, column=41, value=df_def.iat[i,9])
        #100Vケーブルと200Vケーブルの本数を記入
        sheet.cell(row=21, column=42, value=df_def.iat[i,11])
        sheet.cell(row=21, column=48, value=df_def.iat[i,12])

        #特記事項を記入
        sheet.cell(row=33, column=24, value=str(df_def.iat[i,13]))
        print("IP written")
        sheet.cell(row=34, column=24, value=str(df_def.iat[i,14]))
        print("Bios written")
        sheet.cell(row=35, column=24, value=str(df_def.iat[i,15]))
        print("iLO written")
        #sheet['AO12']=df_def.iat[i,2]


#不要な行の削除(HBAがいるときは不要)
sheet.delete_rows(38)
sheet.delete_rows(38)
sheet.delete_rows(38)
    

#編集したExcel workbookの保存
wb.save('sample.xlsx')
#pprint.pprint(list(sheet.values), width=40)
